import time, logging, json, uuid, stripe, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import timedelta
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from email.mime.image import MIMEImage
from django.core.mail import send_mail
from django.contrib.staticfiles import finders
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import EmailMessage
from django.template.loader import get_template
from django.contrib.auth import logout as auth_logout
from django.db import transaction
from django.conf import settings

# IMPORTIAMO SOLO PARTICIPANT E FESTIVALCONFIG
from .models import Participant, FestivalConfig
from .forms import ParticipantForm

stripe.api_key = settings.STRIPE_SECRET_KEY
logger = logging.getLogger(__name__)

def is_admin(user):
    return user.is_staff

def is_organizer(user):
    return user.groups.filter(name="Organizer").exists() or user.is_staff

def get_current_ticket_price():
    return 1025

def is_event_full():
    time_threshold = timezone.now() - timedelta(minutes=30)
    firmly_booked = Participant.objects.filter(status__in=['approved', 'checked_in']).count()
    recently_approved = Participant.objects.filter(status='pending', created_at__gte=time_threshold).count()
    total_count = firmly_booked + recently_approved
    return total_count >= 300

def event_full(request):
    return render(request, 'festival/event_full.html')

def home(request):
    if is_event_full():
        return render(request, 'festival/event_full.html')
    else:
        return render(request, 'festival/home.html')

def about(request):
    return render(request, 'festival/about.html')

def statuto_view(request):
    return render(request, 'festival/statuto.html')

def privacy_view(request):
    return render(request, 'festival/privacy.html')

def payment_cancel(request):
    return render(request, 'festival/payment/payment_cancel.html')

# --- VISTA REGISTRAZIONE AGGIORNATA (SINGOLO UTENTE) ---
def register(request):
    if is_event_full():
        return render(request, 'festival/event_full.html')

    current_price_centesimi = get_current_ticket_price()

    if request.method == 'POST':
        form = ParticipantForm(request.POST)

        if form.is_valid():
            try:
                with transaction.atomic():
                    participant = form.save(commit=False)
                    participant.status = 'pending'
                    participant.save() # Ottiene UUID e salva nel DB

                    # RIMOSSO participant.generate_qr_code() da qui. 
                    # Lo generiamo solo se paga per non intasare il server!

                    base_url = request.build_absolute_uri('/')[:-1]
                    gateway = request.POST.get('gateway', 'stripe')

                    if gateway == 'paypal':
                        messages.warning(request, "Il pagamento con PayPal sarà disponibile a breve.")
                        return redirect('register')
                        
                    elif gateway == 'stripe':
                        expiration_time = int(time.time()) + (30 * 60)
                        
                        checkout_session = stripe.checkout.Session.create(
                            customer_email=participant.email,
                            line_items=[{
                                'price_data': {
                                    'currency': 'eur',
                                    'product_data': {
                                        'name': 'Tesseramento e Ingresso Flanella Fest',
                                    },
                                    'unit_amount': current_price_centesimi,
                                },
                                'quantity': 1,
                            }],
                            mode='payment',
                            expires_at=expiration_time,
                            metadata={
                                'participant_id': str(participant.unique_id),
                                'is_test_payment': 'false'
                            },
                            success_url=base_url + '/register/success/?session_id={CHECKOUT_SESSION_ID}',
                            cancel_url=base_url + '/register/cancel/',
                        )
                        return redirect(checkout_session.url, code=303)

            except Exception as e:
                messages.error(request, f"Errore con il sistema di pagamento: {str(e)}")
                return render(request, 'festival/payment/register.html', {
                    'form': form, 'ticket_price': current_price_centesimi / 100 
                })
    else:
        form = ParticipantForm()
    
    return render(request, 'festival/payment/register.html', {
        'form': form, 'ticket_price': current_price_centesimi / 100
    })

def payment_success(request):
    session_id = request.GET.get('session_id')
    participant = None
    
    if session_id:
        try:
            session = stripe.checkout.Session.retrieve(session_id)
            participant_id = session.metadata.get('participant_id')
            
            if participant_id:
                participant = Participant.objects.get(unique_id=participant_id)
                
                # FALLBACK DI SICUREZZA AGGIORNATO
                if participant.status == 'pending':
                    participant.status = 'approved'
                    participant.pagato = True
                    participant.metodo_pagamento = 'stripe'
                    
                    if not participant.qr_code:
                        participant.generate_qr_code()
                    
                    participant.save()
                    
                    # Invio email con log
                    send_approval_email(participant)
                    
        except Exception as e:
            print(f"Errore nel recupero sessione Stripe: {str(e)}")
            
    return render(request, 'festival/payment/payment_success.html', {'participant': participant})

# --- VISTE ADMIN / STAFF ---
# @login_required
# @user_passes_test(is_organizer)
# def approved_requests(request):
#     # Non essendoci più gli ordini raggruppati, mostriamo solo i soci pagati
#     participants_approved = Participant.objects.filter(status='approved').order_by('-created_at')
#     users_approved = participants_approved.count()
#     users_checked_in = Participant.objects.filter(status='checked_in').count()

#     return render(request, 'festival/approved_requests.html', {
#         'participants': participants_approved,
#         'users_checked_in': users_checked_in,
#         'users_approved': users_approved
#     })

@login_required
@user_passes_test(is_admin)
def participant_detail(request, pk):
    participant = get_object_or_404(Participant, pk=pk)
    return render(request, 'festival/participant_detail.html', {'participant': participant})

def send_approval_email(participant):
    try:
        subject = 'Tesseramento e Registrazione approvati - Flanella Fest'
        html_message = render_to_string('festival/approval_email.html', {'participant': participant})
        
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[participant.email]
        )
        email.content_subtype = 'html'

        # --- 1. Allegato Banner Header via CID ---
        banner_path = finders.find('img/elmo.png')
        if banner_path:
            with open(banner_path, 'rb') as f:
                banner_img = MIMEImage(f.read())
                # Assegniamo l'ID che richiameremo nell'HTML
                banner_img.add_header('Content-ID', '<banner_header>')
                banner_img.add_header('Content-Disposition', 'inline', filename='banner.jpeg')
                email.attach(banner_img)

        # --- 2. Allegato QR Code ---
        if participant.qr_code and hasattr(participant.qr_code, 'path'):
            try:
                email.attach_file(participant.qr_code.path)
            except Exception as e:
                logger.error(f"Impossibile allegare QR Code: {str(e)}")

        email.send(fail_silently=False)
        return True

    except Exception as e:
        logger.error(f"Errore invio mail a {participant.email}: {str(e)}")
        return False

def send_payment_failed_email(participant):
    subject = 'Aggiornamento registrazione - Flanella Fest'
    html_message = render_to_string('festival/rejection_email.html', {'participant': participant})
    email = EmailMessage(subject=subject, body=html_message, from_email=settings.DEFAULT_FROM_EMAIL, to=[participant.email])
    email.content_subtype = 'html'
    try:
        email.send()
    except Exception:
        pass


# --- CONTROLLO QR ---
@login_required
@user_passes_test(is_organizer)
def check_qr(request):
    n_approved = Participant.objects.filter(status='approved').count()
    n_checked_in = Participant.objects.filter(status='checked_in').count()
    result = None
    if request.method == 'POST':
        qr_uuid = request.POST.get('qr_uuid')
        if qr_uuid:
            class DummyRequest:
                method = 'POST'
                body = json.dumps({'qr_data': qr_uuid}).encode()
            response = check_qr_result(DummyRequest())
            result = json.loads(response.content)
            request.session['manual_qr_result'] = result
            return HttpResponseRedirect(reverse('check_qr'))

    if 'manual_qr_result' in request.session:
        result = request.session['manual_qr_result']
        del request.session['manual_qr_result']

    return render(request, 'festival/check_qr.html', {
        'n_checked_in': n_checked_in,
        'total_approved': n_approved + n_checked_in,
        'result': result
    })

@csrf_exempt
def check_qr_result(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            qr_data = data.get('qr_data')
            if not qr_data:
                return JsonResponse({'valid': False, 'status': 'Dati QR Code mancanti o non validi.'})

            qr_data = qr_data.strip()
            try:
                try:
                    uuid_obj = uuid.UUID(qr_data)
                    participant = Participant.objects.get(unique_id=uuid_obj)
                except (ValueError, TypeError):
                    participant = Participant.objects.get(unique_id=qr_data)

                name = f"{participant.first_name} {participant.last_name}"
                if participant.status == 'approved':
                    participant.status = 'checked_in'
                    participant.save()
                    return JsonResponse({'valid': True, 'name': name, 'status': 'Benvenuto! Partecipante verificato.'})
                elif participant.status == 'checked_in':
                    return JsonResponse({'valid': False, 'name': name, 'status': 'Questo QR code è già stato usato.'})
                else:
                    return JsonResponse({'valid': False, 'name': name, 'status': f'Stato: {participant.get_status_display()}'})
            except Participant.DoesNotExist:
                return JsonResponse({'valid': False, 'status': f'QR code non valido o non trovato: {qr_data}'})
            except Exception as e:
                return JsonResponse({'valid': False, 'status': f'Errore durante la verifica: {str(e)}'})

        except json.JSONDecodeError:
            return JsonResponse({'valid': False, 'status': 'Errore nella decodifica dei dati JSON.'})
        except Exception as e:
            return JsonResponse({'valid': False, 'status': f'Errore interno: {str(e)}'})

    return JsonResponse({'valid': False, 'status': 'Metodo non valido. Usa POST.'})


def approved_partecipants(request):
    approved_partecipants = Participant.objects.filter(status='approved')
    return render(request, 'festival/approved_partecipants.html', {'participants': approved_partecipants})

@login_required
@user_passes_test(is_organizer)
def checked_in_partecipants(request):
    checked_in_partecipants = Participant.objects.filter(status='checked_in')
    n_checked_in = checked_in_partecipants.count()
    return render(request, 'festival/checked_in_participants.html', {
        'participants': checked_in_partecipants, 'n_checked_in': n_checked_in
    })

@login_required
@user_passes_test(is_organizer)
def all_participants(request):
    participants = Participant.objects.all().order_by('-created_at')
    n_checked_in = Participant.objects.filter(status='checked_in').count()
    return render(request, 'festival/all_participants.html', { 'participants' : participants, 'n_checked_in': n_checked_in })

# @login_required
# @user_passes_test(is_organizer)
# def all_participants_pdf(request):
#     from weasyprint import HTML
#     participants = Participant.objects.filter(status__in=['approved', 'checked_in'])
#     n_checked_in = Participant.objects.filter(status='checked_in').count()
#     template = get_template('festival/all_participants_pdf.html')
#     html_string = template.render({'participants' : participants, 'n_checked_in': n_checked_in })
#     base_url = request.build_absolute_uri('/')
#     pdf_file = HTML(string=html_string, base_url=base_url).write_pdf()
#     response = HttpResponse(pdf_file, content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment;filename="partecipanti.pdf"'
#     return response

@login_required
@user_passes_test(is_organizer)
def export_participants_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Libro Soci & Partecipanti"

    # 1. TUTTI I CAMPI DEL FORM E DEL SISTEMA
    headers = [
        "ID Unico",
        "Nome",
        "Cognome",
        "Data di Nascita",
        "Luogo di Nascita",
        "Codice Fiscale",
        "Indirizzo",
        "Città",
        "CAP",
        "Provincia",
        "Telefono",
        "Email",
        "Metodo Pagamento",
        "Stato Check-in",
        "Data Iscrizione"
    ]
    sheet.append(headers)

    # Stile Intestazione
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid") # Nero Inchiostro
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2. RECUPERO PARTECIPANTI (Approvati e Checked-in)
    participants = Participant.objects.filter(status__in=['approved', 'checked_in']).order_by('-created_at')

    # 3. POPOLAMENTO RIGHE CON OGNI SINGOLO DATO
    for p in participants:
        sheet.append([
            str(p.unique_id),
            p.first_name,
            p.last_name,
            p.data_di_nascita.strftime("%d/%m/%Y") if p.data_di_nascita else "",
            p.luogo_di_nascita or "",
            p.codice_fiscale or "",
            p.indirizzo or "",
            p.citta or "",
            p.cap or "",
            p.provincia or "",
            p.phone or "",
            p.email or "",
            p.get_metodo_pagamento_display() if hasattr(p, 'get_metodo_pagamento_display') else p.metodo_pagamento,
            "Arrivato" if p.status == 'checked_in' else "In attesa",
            p.created_at.strftime("%d/%m/%Y %H:%M")
        ])

    # 4. AUTO-RIDIMENSIONAMENTO DELLE COLONNE
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max(max_length + 3, 12)
        sheet.column_dimensions[column].width = adjusted_width

    # 5. RISPOSTA ED ESPORTAZIONE FILE
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="libro_soci_flanellafest.xlsx"'
    
    workbook.save(response)
    return response

@login_required
@user_passes_test(is_organizer)
def logout(request):
    auth_logout(request)
    return redirect('home')


# --- WEBHOOK STRIPE AGGIORNATO PER IL SINGOLO PARTECIPANTE ---
@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    event = None

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, settings.STRIPE_WEBHOOK_SECRET
        )
    except (ValueError, stripe.error.SignatureVerificationError):
        return HttpResponse(status=400)

    if event['type'] == 'checkout.session.completed':
        session = event['data']['object']
        
        is_test_payment = session.get('metadata', {}).get('is_test_payment', 'false')
        if is_test_payment == 'true':
            send_mail(
                'Stripe Prod Test OK - Flanella Fest',
                'Il test da 50 centesimi è andato a buon fine.',
                settings.DEFAULT_FROM_EMAIL, [settings.EMAIL_HOST_USER], fail_silently=True
            )
            return HttpResponse(status=200)

        participant_id = session.get('metadata', {}).get('participant_id')
        
        if participant_id:
            try:
                participant = Participant.objects.get(unique_id=participant_id)
                # Se è ancora in pending, lo approviamo
                if participant.status == 'pending':
                    participant.status = 'approved'
                    participant.pagato = True
                    participant.metodo_pagamento = 'stripe'
                    
                    if not participant.qr_code:
                        participant.generate_qr_code()
                    participant.save()
                    
                    send_approval_email(participant)
            except Participant.DoesNotExist:
                return HttpResponse(status=404)

    elif event['type'] == 'checkout.session.expired':
        session = event['data']['object']
        if session.get('metadata', {}).get('is_test_payment') == 'true':
            return HttpResponse(status=200)

        participant_id = session.get('metadata', {}).get('participant_id')
        if participant_id:
            try:
                participant = Participant.objects.get(unique_id=participant_id)
                if participant.status == 'pending':
                    participant.status = 'rejected'
                    participant.save()
            except Participant.DoesNotExist:
                pass 

    return HttpResponse(status=200)

@login_required
@user_passes_test(is_admin)
def admin_management(request):
    config = FestivalConfig.objects.first()
    if not config:
        config = FestivalConfig.objects.create(ticket_price_centesimi=settings.TICKET_PRICE_CENTESIMI)

    if request.method == 'POST':
        new_price_euro = request.POST.get('ticket_price')
        if new_price_euro:
            try:
                centesimi = int(float(new_price_euro.replace(',', '.')) * 100)
                config.ticket_price_centesimi = centesimi
                config.save()
                messages.success(request, f"Prezzo aggiornato con successo a {new_price_euro} €!")
            except ValueError:
                messages.error(request, "Inserisci un valore numerico valido.")
            return redirect('admin_management')

    return render(request, 'festival/admin_management.html', {
        'current_price_euro': config.ticket_price_centesimi / 100
    })

@login_required
@user_passes_test(is_admin)
def stripe_test_checkout(request):
    base_url = request.build_absolute_uri('/')[:-1]
    try:
        checkout_session = stripe.checkout.Session.create(
            customer_email=request.user.email,
            line_items=[{
                'price_data': {
                    'currency': 'eur',
                    'product_data': {'name': 'TEST SISTEMA DI PAGAMENTO (50 CENTESIMI)'},
                    'unit_amount': 50, 
                },
                'quantity': 1,
            }],
            mode='payment',
            metadata={'is_test_payment': 'true'},
            success_url = base_url + '/register/success/?session_id={CHECKOUT_SESSION_ID}',
            cancel_url=base_url + '/register/cancel/',
        )
        return redirect(checkout_session.url, code=303)
    except Exception as e:
        messages.error(request, f"Errore test Stripe: {str(e)}")
        return redirect('admin_management')